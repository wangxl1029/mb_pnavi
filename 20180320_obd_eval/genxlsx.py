#!/usr/bin/python3
# -*- coding=utf-8 -*-

#standard lib
import os
import io
import sys
import getopt
import json
from configparser import ConfigParser
#3rd party lib
import pymongo
from openpyxl import Workbook
from openpyxl import load_workbook

config = ConfigParser()
db = None
db_client = None

def db_open_once():
	global db, db_client
	if None == db_client:
		db_host = (config['mongo']['host_addr'], int(config['mongo']['host_port']))
		db_client = pymongo.MongoClient(db_host[0], db_host[1])
	if None == db:
		db = db_client.get_database(config['mongo']['database'])
def db_close():
	global db, db_client
	if not None == db:
		db = None
	if not None == db_client:
		db_client.close()
		db_client = None

def createExcelTemplate(xlsPath, tabRows):
	wb = Workbook()
	ws = wb.active
	ws.title = 'Learning Outcome'
	rowHeader = ['Obj ID', 'INFERENCE', 'position info', 'remarks']
	ws.append(rowHeader)
	for row in tabRows:
		ws.append(row)
		#print(tabRows)
	wb.create_sheet('Map View')
	wb.save(xlsPath)

def extractJourneyInfo(celTag, jsnDir, strGuid):
	db_open_once()
	jn = db.get_collection(config['mongo']['colle_journeys'])
	JOURNEY_MAX = 3
	jn_idx = 0
	for cur_jn in jn.find({'guid': strGuid}, \
			{'guid': 1, 'start_lat': 1, 'start_lon': 1, 'end_lat': 1, 'end_lon': 1, 'via_points': 1}):
		if jn_idx < JOURNEY_MAX: pass
		else: break
		obj_id = str(cur_jn['_id'])
		jsnPath = os.path.join(jsnDir, f'{celTag}-J{jn_idx}_obj_{obj_id}.json')
		#print(jsnPath)
		if os.path.exists(jsnPath):
			print(f'\"{jsnPath}\" already exists!')
		else:
			cur_jn['_id'] = obj_id
			jsnFile = open(jsnPath, "w", encoding="utf-8")
			jsnFile.write(json.dumps(cur_jn, indent=4))
			jsnFile.close()
			print(f'\"{jsnPath}\" done.')
		jn_idx += 1
	return

def extractLearningOutcome(celTag, jsnDir, strGuid):
	db_open_once()
	#global db
	lo = db.get_collection(config['mongo']['collo_learning_outcomes'])
	#print(strGuid)
	tabRows = []
	loIdx = 0
	for cur_lo in lo.find({'guid': strGuid}, {'guid':1, 'type':1, 'latitude': 1, 'longitude': 1}):
		#print(curitm['_id'], curitm['type'])
		obj_id = str(cur_lo['_id'])
		tabRows += [[obj_id, cur_lo['type']]]
		loPath = os.path.join(jsnDir, f'{celTag}-LO#{loIdx}_obj_{obj_id}.json')
		#print(loPath)
		if os.path.exists(loPath):
			print(f'\"{loPath}\" already exists!')
		else:
			cur_lo['_id'] = obj_id
			loFile = open(loPath, "w", encoding="utf-8")
			loFile.write(json.dumps(cur_lo, indent=4))
			loFile.close()
		loIdx += 1
	return tabRows

def main(iniFile):
	if config.read(iniFile) == []:
		print(f'Init file \"{iniFile}\" read error!');
		return
	wb = load_workbook(config['excel.workbook']['obd_test'])
	ws = wb[config['excel.wk_obd_test.sheet']['homelist']]
	dirNum = 0
	xlsNum = 0
	guid_start_rowno = int(config['excel.wk_obd_test.ws_homelist']['guid_start_rowno'])
	guid_end_rowno = int(config['excel.wk_obd_test.ws_homelist']['guid_end_rowno'])
	for rownum in range(guid_start_rowno, guid_end_rowno + 1):
		xyTag = 'A'+ str(rownum)
		xyGUID = 'C' + str(rownum)
		celTag = ws[xyTag].value
		celGuid = ws[xyGUID].value
		dirPath = '{}_guid_{}'.format(celTag, celGuid)
		#print(dirName)
		if not os.path.exists(dirPath):
			os.makedirs(dirPath)
			dirNum += 1
		else:
			print(f'{dirPath} already exist!')
		extractJourneyInfo(celTag, dirPath, str(celGuid))
		#Learning outcome excel
		xlsPath = '{}_guid_{}_templ.xlsx'.format(celTag, celGuid)
		if not os.path.exists(xlsPath):
			tabRows = extractLearningOutcome(celTag, dirPath, str(celGuid))
			createExcelTemplate(xlsPath, tabRows)
			xlsNum += 1
		else:
			print(f'{xlsPath} already exist!')
		
	print(f'{dirNum} folder(s) created!')
	print(f'{xlsNum} excel template(s) created!')
	db_close()

def DispUsage():
	print('''Usage!
	-h	--help
	-i	--initfile=somefile
	''')

if __name__ == '__main__':
	if len(sys.argv) == 1:
		DispUsage()
		sys.exit(1)
	try: 
		opts, args = getopt.getopt(sys.argv[1:], 'hi:', ['help', 'initfile='])
	except getopt.GetoptError as err:
		print(err)
		DispUsage()
		sys.exit(2)
	for o, a in opts:
		if o in ("-h", "--help"):
			DispUsage()
		elif o in ('-i', '--initfile'):
			main(iniFile=a)
		else:
			DispUsage()

